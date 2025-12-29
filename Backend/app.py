from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from pymongo import MongoClient, DESCENDING
from bson import ObjectId
from datetime import datetime, timedelta
import os
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from urllib.parse import quote_plus
import jwt
from functools import wraps

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)

# MongoDB Configuration with auto URL encoding
MONGO_USER = os.getenv("MONGO_USER", "JobPortal")
MONGO_PASSWORD = os.getenv("MONGO_PASSWORD", "")
MONGO_CLUSTER = os.getenv("MONGO_CLUSTER", "jobportal.vrdkavz.mongodb.net")

# Admin Credentials
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")
SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-key")

# Auto-encode credentials
username = quote_plus(MONGO_USER)
password = quote_plus(MONGO_PASSWORD)

# Build connection string
MONGO_URI = f"mongodb+srv://{username}:{password}@{MONGO_CLUSTER}/job_portal?retryWrites=true&w=majority"

client = MongoClient(MONGO_URI)
db = client['job_portal']
applications_collection = db['applications']
jobs_collection = db['jobs']

# File Upload Configuration
UPLOAD_FOLDER = "uploads/resumes"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['SECRET_KEY'] = SECRET_KEY

# JWT Token decorator
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({"error": "Token is missing"}), 401
        
        try:
            # Remove 'Bearer ' prefix if present
            if token.startswith('Bearer '):
                token = token[7:]
            
            data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token has expired"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "Invalid token"}), 401
        
        return f(*args, **kwargs)
    
    return decorated

# Helper function to serialize MongoDB documents
def serialize_doc(doc):
    """Convert MongoDB document to JSON-serializable dict"""
    if doc is None:
        return None
    doc['_id'] = str(doc['_id'])
    if 'ID' not in doc and 'id' not in doc:
        doc['id'] = str(doc['_id'])
    return doc

# ============ PUBLIC ROUTES (No Authentication) ============

@app.route("/health", methods=["GET"])
def health():
    try:
        client.server_info()
        app_count = applications_collection.count_documents({})
        job_count = jobs_collection.count_documents({})
        return jsonify({
            "status": "ok",
            "message": "Backend is running",
            "database": "connected",
            "applications_count": app_count,
            "jobs_count": job_count
        }), 200
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": "Database connection failed",
            "error": str(e)
        }), 500

@app.route("/apply", methods=["POST"])
def apply_job():
    try:
        form = request.form
        resume = request.files.get("resume")
        
        if not resume:
            return jsonify({"error": "Resume file is required"}), 400

        # Save resume file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{resume.filename}"
        resume_path = os.path.join(UPLOAD_FOLDER, filename)
        resume.save(resume_path)

        # Get next ID (sequential)
        last_app = applications_collection.find_one(sort=[("ID", DESCENDING)])
        next_id = (last_app['ID'] + 1) if last_app and 'ID' in last_app else 1

        # Create application document
        application = {
            "ID": next_id,
            "Position": form.get("position", "N/A"),
            "Name": form.get("name"),
            "Email": form.get("email"),
            "Phone": form.get("phone"),
            "College": form.get("college"),
            "Degree": form.get("degree"),
            "Passout Year": form.get("year"),
            "Skills": form.get("skills"),
            "Resume File": filename,
            "Resume Path": resume_path,
            "Applied Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Status": "Pending",
            "Created At": datetime.now()
        }

        # Insert into MongoDB
        result = applications_collection.insert_one(application)
        
        return jsonify({
            "message": "Application submitted successfully",
            "id": str(result.inserted_id)
        }), 201
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ============ JOB ROUTES (PUBLIC) ============

@app.route("/jobs", methods=["GET"])
def get_jobs():
    """Public route - Get all active jobs for career portal"""
    try:
        query = {"status": "Active"}
        
        # Optional filters
        job_type = request.args.get('type')
        experience = request.args.get('experience')
        
        if job_type and job_type != 'all':
            query['type'] = job_type
        
        if experience and experience != 'all':
            query['experience'] = experience
        
        jobs = list(jobs_collection.find(query).sort("created_at", DESCENDING))
        jobs = [serialize_doc(job) for job in jobs]
        
        return jsonify(jobs), 200
        
    except Exception as e:
        print(f"Error fetching jobs: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/jobs/<int:job_id>", methods=["GET"])
def get_job_by_id(job_id):
    """Public route - Get single job details"""
    try:
        job = jobs_collection.find_one({"id": job_id, "status": "Active"})
        
        if not job:
            return jsonify({"error": "Job not found"}), 404
        
        job = serialize_doc(job)
        return jsonify(job), 200
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ============ ADMIN LOGIN ROUTE ============

@app.route("/admin/login", methods=["POST"])
def admin_login():
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({"error": "Username and password required"}), 400
        
        # Check credentials
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            # Generate JWT token (expires in 24 hours)
            token = jwt.encode({
                'username': username,
                'exp': datetime.utcnow() + timedelta(hours=24)
            }, SECRET_KEY, algorithm="HS256")
            
            return jsonify({
                "message": "Login successful",
                "token": token,
                "username": username
            }), 200
        else:
            return jsonify({"error": "Invalid credentials"}), 401
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ============ PROTECTED ADMIN ROUTES - APPLICATIONS ============

@app.route("/admin/applications", methods=["GET"])
@token_required
def get_applications():
    try:
        query = {}
        
        position = request.args.get('position')
        status = request.args.get('status')
        
        if position and position != 'all':
            query['Position'] = position
        
        if status and status != 'all':
            query['Status'] = status
        
        applications = list(applications_collection.find(query).sort("ID", DESCENDING))
        applications = [serialize_doc(app) for app in applications]
        
        return jsonify(applications), 200
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/statistics", methods=["GET"])
@token_required
def get_statistics():
    try:
        total = applications_collection.count_documents({})
        
        position_pipeline = [
            {"$group": {"_id": "$Position", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        by_position = {item['_id']: item['count'] 
                      for item in applications_collection.aggregate(position_pipeline)}
        
        status_pipeline = [
            {"$group": {"_id": "$Status", "count": {"$sum": 1}}}
        ]
        by_status_result = {item['_id']: item['count'] 
                           for item in applications_collection.aggregate(status_pipeline)}
        
        by_status = {
            "Pending": by_status_result.get("Pending", 0),
            "Reviewed": by_status_result.get("Reviewed", 0),
            "Shortlisted": by_status_result.get("Shortlisted", 0),
            "Rejected": by_status_result.get("Rejected", 0)
        }
        
        recent = list(applications_collection.find()
                     .sort("Created At", DESCENDING)
                     .limit(5))
        recent = [serialize_doc(app) for app in recent]
        
        stats = {
            "total": total,
            "by_position": by_position,
            "by_status": by_status,
            "recent_applications": recent
        }
        
        return jsonify(stats), 200
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/application/<int:app_id>/status", methods=["PUT"])
@token_required
def update_status(app_id):
    try:
        data = request.json
        new_status = data.get('status')
        
        if new_status not in ['Pending', 'Reviewed', 'Shortlisted', 'Rejected']:
            return jsonify({"error": "Invalid status"}), 400
        
        result = applications_collection.update_one(
            {"ID": app_id},
            {"$set": {
                "Status": new_status,
                "Updated At": datetime.now()
            }}
        )
        
        if result.matched_count == 0:
            return jsonify({"error": "Application not found"}), 404
        
        return jsonify({"message": "Status updated successfully"}), 200
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/application/<int:app_id>", methods=["DELETE"])
@token_required
def delete_application(app_id):
    try:
        application = applications_collection.find_one({"ID": app_id})
        
        if not application:
            return jsonify({"error": "Application not found"}), 404
        
        resume_path = application.get('Resume Path')
        if resume_path and os.path.exists(resume_path):
            try:
                os.remove(resume_path)
            except Exception as e:
                print(f"Warning: Could not delete resume file: {e}")
        
        applications_collection.delete_one({"ID": app_id})
        
        return jsonify({"message": "Application deleted successfully"}), 200
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/download-excel", methods=["GET"])
@token_required
def admin_download_excel():
    try:
        query = {}
        position = request.args.get('position')
        
        if position and position != 'all':
            query['Position'] = position
        
        applications = list(applications_collection.find(query).sort("ID", DESCENDING))
        
        if not applications:
            return jsonify({"error": "No data available"}), 404
        
        df_data = []
        for app in applications:
            df_data.append({
                "ID": app.get("ID"),
                "Position": app.get("Position"),
                "Name": app.get("Name"),
                "Email": app.get("Email"),
                "Phone": app.get("Phone"),
                "College": app.get("College"),
                "Degree": app.get("Degree"),
                "Passout Year": app.get("Passout Year"),
                "Skills": app.get("Skills"),
                "Resume File": app.get("Resume File"),
                "Applied Date": app.get("Applied Date"),
                "Status": app.get("Status")
            })
        
        df = pd.DataFrame(df_data)
        
        output = BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        
        wb = load_workbook(output)
        ws = wb.active
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ============ RESUME SERVING ROUTES ============

@app.route('/uploads/resumes/<filename>')
def serve_resume(filename):
    """Serve resume files (Protected route - requires authentication)"""
    try:
        # Get token from query params or Authorization header
        token = request.args.get('token') or request.headers.get('Authorization', '').replace('Bearer ', '')
        
        if not token:
            return jsonify({'error': 'Authentication required'}), 401
        
        # Verify the token
        try:
            jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        
        # Security check: prevent directory traversal
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': 'Invalid filename'}), 400
        
        # Check if file exists
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'Resume not found'}), 404
        
        # Send the file
        return send_from_directory(UPLOAD_FOLDER, filename)
        
    except Exception as e:
        print(f"Error serving resume: {str(e)}")
        return jsonify({'error': 'Failed to serve resume'}), 500

@app.route('/admin/download-resume/<filename>')
@token_required
def download_resume(filename):
    """Download resume file (Protected route)"""
    try:
        # Security check: prevent directory traversal
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'error': 'Invalid filename'}), 400
        
        # Check if file exists
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'Resume not found'}), 404
        
        # Send file as attachment
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error downloading resume: {str(e)}")
        return jsonify({'error': 'Failed to download resume'}), 500

# ============ PROTECTED ADMIN ROUTES - JOBS ============

@app.route("/admin/jobs", methods=["GET"])
@token_required
def admin_get_jobs():
    """Admin route - Get all jobs (including inactive)"""
    try:
        jobs = list(jobs_collection.find().sort("created_at", DESCENDING))
        jobs = [serialize_doc(job) for job in jobs]
        return jsonify(jobs), 200
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/jobs", methods=["POST"])
@token_required
def create_job():
    """Create a new job posting"""
    try:
        data = request.json
        
        # Validate required fields
        required_fields = ['title', 'company', 'location', 'type', 'experience', 'salary', 'description']
        for field in required_fields:
            if not data.get(field):
                return jsonify({"error": f"{field} is required"}), 400
        
        # Get next ID
        last_job = jobs_collection.find_one(sort=[("id", DESCENDING)])
        next_id = (last_job['id'] + 1) if last_job and 'id' in last_job else 1
        
        # Create job document
        job = {
            "id": next_id,
            "title": data.get("title"),
            "company": data.get("company"),
            "location": data.get("location"),
            "type": data.get("type"),
            "experience": data.get("experience"),
            "salary": data.get("salary"),
            "postedDate": "Just now",
            "applicants": 0,
            "description": data.get("description"),
            "responsibilities": data.get("responsibilities", []),
            "requirements": data.get("requirements", []),
            "skills": data.get("skills", []),
            "benefits": data.get("benefits", []),
            "status": data.get("status", "Active"),
            "created_at": datetime.now(),
            "updated_at": datetime.now()
        }
        
        result = jobs_collection.insert_one(job)
        
        return jsonify({
            "message": "Job created successfully",
            "id": str(result.inserted_id),
            "job_id": next_id
        }), 201
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/jobs/<int:job_id>", methods=["PUT"])
@token_required
def update_job(job_id):
    """Update an existing job"""
    try:
        data = request.json
        
        update_data = {
            "title": data.get("title"),
            "company": data.get("company"),
            "location": data.get("location"),
            "type": data.get("type"),
            "experience": data.get("experience"),
            "salary": data.get("salary"),
            "description": data.get("description"),
            "responsibilities": data.get("responsibilities", []),
            "requirements": data.get("requirements", []),
            "skills": data.get("skills", []),
            "benefits": data.get("benefits", []),
            "status": data.get("status", "Active"),
            "updated_at": datetime.now()
        }
        
        result = jobs_collection.update_one(
            {"id": job_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            return jsonify({"error": "Job not found"}), 404
        
        return jsonify({"message": "Job updated successfully"}), 200
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/jobs/<int:job_id>", methods=["DELETE"])
@token_required
def delete_job(job_id):
    """Delete a job posting"""
    try:
        result = jobs_collection.delete_one({"id": job_id})
        
        if result.deleted_count == 0:
            return jsonify({"error": "Job not found"}), 404
        
        return jsonify({"message": "Job deleted successfully"}), 200
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/jobs/<int:job_id>/toggle-status", methods=["PUT"])
@token_required
def toggle_job_status(job_id):
    """Toggle job status between Active and Inactive"""
    try:
        job = jobs_collection.find_one({"id": job_id})
        
        if not job:
            return jsonify({"error": "Job not found"}), 404
        
        new_status = "Inactive" if job.get("status") == "Active" else "Active"
        
        jobs_collection.update_one(
            {"id": job_id},
            {"$set": {"status": new_status, "updated_at": datetime.now()}}
        )
        
        return jsonify({"message": "Status updated", "new_status": new_status}), 200
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ============ HELPER FUNCTIONS ============

def create_indexes():
    """Create MongoDB indexes for optimized queries"""
    try:
        # Applications indexes
        applications_collection.create_index("ID", unique=True)
        applications_collection.create_index("Email")
        applications_collection.create_index("Position")
        applications_collection.create_index("Status")
        applications_collection.create_index([("Created At", DESCENDING)])
        
        # Jobs indexes
        jobs_collection.create_index("id", unique=True)
        jobs_collection.create_index("status")
        jobs_collection.create_index("title")
        jobs_collection.create_index([("created_at", DESCENDING)])
        
        print("[OK] MongoDB indexes created successfully")
    except Exception as e:
        print(f"[WARNING] Could not create indexes: {e}")

# ============ MAIN ============

if __name__ == "__main__":
    print("\n" + "="*60)
    print("   JOB PORTAL BACKEND - MONGODB VERSION")
    print("="*60)
    print(f"\n[*] Connecting to MongoDB...")
    
    try:
        client.server_info()
        print("[OK] MongoDB connection successful!")
        print(f"[DB] Database: job_portal")
        print(f"[DB] Collections: applications, jobs")
        print(f"[ADMIN] Admin Username: {ADMIN_USERNAME}")
        print(f"[UPLOADS] Resume folder: {UPLOAD_FOLDER}")
        
        create_indexes()
        
        print("\n[*] Starting Flask server...")
        print("="*60 + "\n")
        
        app.run(debug=True, port=5000, host='0.0.0.0')
        
    except Exception as e:
        print(f"\n[ERROR] Failed to connect to MongoDB!")
        print(f"Error: {e}")
        print("\n")